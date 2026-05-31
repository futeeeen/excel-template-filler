from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Callable

from openpyxl import load_workbook


DEFAULT_PLACEHOLDER = "BX-XXX_FSD議題名稱"
INVALID_FILENAME_CHARS = r'<>:"/\|?*'


@dataclass(frozen=True)
class ReplacementTarget:
    sheet_name: str
    cell: str


@dataclass(frozen=True)
class FillOptions:
    source_file: Path
    template_file: Path
    source_sheet: str
    template_sheet: str = ""
    source_column: str = "E"
    start_row: int = 2
    target_cell: str = "A1"
    placeholder: str = DEFAULT_PLACEHOLDER
    output_base_dir: Path | None = None
    output_prefix: str = "output_checklists"
    replacement_targets: list[ReplacementTarget] | None = None
    replace_workbook: bool = True
    replace_filename: bool = True


@dataclass(frozen=True)
class FillResult:
    output_dir: Path
    created_files: list[Path]
    skipped_duplicates: int


class ValidationError(Exception):
    pass


def get_workbook_sheets(path: Path) -> list[str]:
    if not path.exists():
        raise ValidationError(f"找不到檔案：{path}")
    if path.suffix.lower() != ".xlsx":
        raise ValidationError("目前只支援 .xlsx 檔案。")
    wb = load_workbook(path, read_only=True, data_only=True)
    return list(wb.sheetnames)


def normalize_column(column: str) -> str:
    value = column.strip().upper()
    if not re.fullmatch(r"[A-Z]{1,3}", value):
        raise ValidationError("來源欄位請輸入英文字母，例如 E 或 G。")
    return value


def read_names(source_file: Path, sheet_name: str, column: str, start_row: int) -> list[str]:
    if start_row < 1:
        raise ValidationError("起始列必須大於 0。")

    column = normalize_column(column)
    wb = load_workbook(source_file, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValidationError(f"來源檔找不到工作表：{sheet_name}")

    ws = wb[sheet_name]
    names: list[str] = []
    for row in range(start_row, ws.max_row + 1):
        value = ws[f"{column}{row}"].value
        if value is None:
            continue
        name = str(value).strip()
        if name:
            names.append(name)
    return names


def get_replacement_targets(options: FillOptions) -> list[ReplacementTarget]:
    if options.replacement_targets is not None:
        return [
            ReplacementTarget(target.sheet_name.strip(), target.cell.strip().upper())
            for target in options.replacement_targets
            if target.sheet_name.strip() and target.cell.strip()
        ]
    if options.template_sheet.strip() and options.target_cell.strip():
        return [ReplacementTarget(options.template_sheet.strip(), options.target_cell.strip().upper())]
    return []


def validate_options(options: FillOptions) -> None:
    if not options.source_file.exists():
        raise ValidationError("請選擇存在的來源清單 Excel。")
    if not options.template_file.exists():
        raise ValidationError("請選擇存在的範本 Excel。")
    if options.source_file.suffix.lower() != ".xlsx" or options.template_file.suffix.lower() != ".xlsx":
        raise ValidationError("來源清單與範本都必須是 .xlsx 檔案。")
    if not options.source_sheet.strip():
        raise ValidationError("請選擇來源清單的工作表。")
    if not options.placeholder.strip():
        raise ValidationError("替換文字不可空白。")
    if not options.replace_workbook and not options.replace_filename:
        raise ValidationError("請至少勾選替換範本內容或替換檔名。")

    normalize_column(options.source_column)
    source_sheets = get_workbook_sheets(options.source_file)
    template_sheets = get_workbook_sheets(options.template_file)
    if options.source_sheet not in source_sheets:
        raise ValidationError(f"來源檔找不到工作表：{options.source_sheet}")

    names = read_names(
        options.source_file,
        options.source_sheet,
        options.source_column,
        options.start_row,
    )
    if not names:
        raise ValidationError("來源欄位沒有可產出的名稱，請確認工作表、欄位與起始列。")

    targets = get_replacement_targets(options)
    if options.replace_workbook and not targets:
        raise ValidationError("請至少設定一組要替換的範本工作表與儲存格。")

    if options.replace_workbook:
        wb = load_workbook(options.template_file, read_only=True, data_only=True)
        for target in targets:
            if target.sheet_name not in template_sheets:
                raise ValidationError(f"範本檔找不到工作表：{target.sheet_name}")
            ws = wb[target.sheet_name]
            target_value = ws[target.cell].value
            if target_value is None or options.placeholder not in str(target_value):
                raise ValidationError(
                    f"範本 {target.sheet_name}!{target.cell} 找不到「{options.placeholder}」。"
                )


def sanitize_filename_part(value: str) -> str:
    table = str.maketrans({char: "_" for char in INVALID_FILENAME_CHARS})
    cleaned = value.translate(table).strip().rstrip(".")
    return cleaned or "未命名"


def replace_template_filename(template_name: str, placeholder: str, name: str) -> str:
    safe_name = sanitize_filename_part(name)
    candidates = [placeholder, placeholder.replace("_", " ")]
    result = template_name
    for candidate in candidates:
        if candidate in result:
            result = result.replace(candidate, safe_name)
    if result == template_name:
        stem = Path(template_name).stem
        suffix = Path(template_name).suffix
        result = f"{stem}_{safe_name}{suffix}"
    return result


def unique_path(path: Path) -> tuple[Path, bool]:
    if not path.exists():
        return path, False
    stem = path.stem
    suffix = path.suffix
    for index in range(2, 10000):
        candidate = path.with_name(f"{stem}_{index}{suffix}")
        if not candidate.exists():
            return candidate, True
    raise ValidationError(f"無法建立不重複檔名：{path.name}")


def build_output_dir(options: FillOptions) -> Path:
    base_dir = options.output_base_dir or options.source_file.parent
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    prefix = sanitize_filename_part(options.output_prefix.strip() or "output_checklists")
    return base_dir / f"{prefix}_{timestamp}"


def build_plain_filename(template_file: Path, index: int) -> str:
    return f"{template_file.stem}_{index:03d}{template_file.suffix}"


def fill_templates(
    options: FillOptions,
    progress: Callable[[str], None] | None = None,
) -> FillResult:
    validate_options(options)

    names = read_names(options.source_file, options.source_sheet, options.source_column, options.start_row)
    targets = get_replacement_targets(options)
    output_dir = build_output_dir(options)
    output_dir.mkdir(parents=True, exist_ok=False)

    created_files: list[Path] = []
    skipped_duplicates = 0
    for index, name in enumerate(names, start=1):
        wb = load_workbook(options.template_file)

        if options.replace_workbook:
            for target in targets:
                if target.sheet_name not in wb.sheetnames:
                    raise ValidationError(f"範本檔找不到工作表：{target.sheet_name}")
                ws = wb[target.sheet_name]
                target_value = str(ws[target.cell].value)
                ws[target.cell] = target_value.replace(options.placeholder, name)

        if options.replace_filename:
            new_filename = replace_template_filename(options.template_file.name, options.placeholder, name)
        else:
            new_filename = build_plain_filename(options.template_file, index)

        new_path, duplicated = unique_path(output_dir / new_filename)
        if duplicated:
            skipped_duplicates += 1

        wb.save(new_path)
        created_files.append(new_path)
        if progress:
            progress(f"[{index}/{len(names)}] 已產出：{new_path.name}")

    return FillResult(
        output_dir=output_dir,
        created_files=created_files,
        skipped_duplicates=skipped_duplicates,
    )
