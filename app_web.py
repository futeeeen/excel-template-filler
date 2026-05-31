from __future__ import annotations

import base64
import io
import json
import os
import shutil
import subprocess
import sys
import tempfile
import threading
import zipfile
from http import HTTPStatus
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import urlparse

from openpyxl import load_workbook

from src.excel_template_filler import (
    DEFAULT_PLACEHOLDER,
    FillOptions,
    ReplacementTarget,
    ValidationError,
    fill_templates,
    read_names,
    sanitize_filename_part,
    validate_options,
)


APP_DIR = Path(getattr(sys, "_MEIPASS", Path(__file__).resolve().parent))
EXAMPLES_DIR = APP_DIR / "examples"


def find_browser() -> Path | None:
    env_paths = [
        os.environ.get("PROGRAMFILES", ""),
        os.environ.get("PROGRAMFILES(X86)", ""),
        os.environ.get("LOCALAPPDATA", ""),
    ]
    candidates = [
        Path(env_paths[0]) / "Google/Chrome/Application/chrome.exe",
        Path(env_paths[1]) / "Google/Chrome/Application/chrome.exe",
        Path(env_paths[2]) / "Google/Chrome/Application/chrome.exe",
        Path(env_paths[0]) / "Microsoft/Edge/Application/msedge.exe",
        Path(env_paths[1]) / "Microsoft/Edge/Application/msedge.exe",
        Path(env_paths[2]) / "Microsoft/Edge/Application/msedge.exe",
    ]
    for candidate in candidates:
        if candidate.exists():
            return candidate
    return None


def open_app_url(url: str) -> None:
    browser = find_browser()
    if browser:
        subprocess.Popen([str(browser), url], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        return
    os.startfile(url)


HTML = r"""<!doctype html>
<html lang="zh-Hant">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Excel Checklist 產生工具</title>
  <style>
    :root {
      --ink: #142033;
      --muted: #627087;
      --line: #d7dde8;
      --soft: #f5f7fa;
      --panel: #ffffff;
      --accent: #0f766e;
      --accent-dark: #0b5f59;
      --danger: #b42318;
      --blue: #1d4ed8;
    }
    * { box-sizing: border-box; }
    body {
      margin: 0;
      font-family: "Microsoft JhengHei UI", "Microsoft JhengHei", "Segoe UI", sans-serif;
      color: var(--ink);
      background: #edf1f6;
    }
    header {
      background: var(--panel);
      border-bottom: 1px solid var(--line);
      padding: 18px 28px 12px;
    }
    h1 { margin: 0 0 12px; font-size: 24px; letter-spacing: 0; }
    .tabs { display: flex; gap: 8px; }
    .tab {
      border: 1px solid var(--line);
      background: #f9fafc;
      padding: 10px 18px;
      border-radius: 6px;
      cursor: pointer;
      font-size: 15px;
    }
    .tab.active {
      background: var(--accent);
      color: #fff;
      border-color: var(--accent);
      font-weight: 700;
    }
    main {
      max-width: 1120px;
      margin: 18px auto 28px;
      padding: 0 18px;
    }
    section.page { display: none; }
    section.page.active { display: block; }
    .shell {
      background: var(--panel);
      border: 1px solid var(--line);
      border-radius: 8px;
      padding: 22px;
      box-shadow: 0 2px 10px rgba(20, 32, 51, .06);
    }
    h2 { margin: 0 0 18px; font-size: 22px; }
    h3 { margin: 0 0 12px; font-size: 17px; }
    p, li { line-height: 1.7; }
    .config-block {
      border: 1px solid var(--line);
      border-radius: 8px;
      padding: 16px;
      margin-bottom: 16px;
      background: #fff;
    }
    .block-title {
      display: flex;
      justify-content: space-between;
      gap: 12px;
      align-items: flex-start;
      margin-bottom: 14px;
    }
    .block-title p { margin: 4px 0 0; color: var(--muted); font-size: 14px; }
    .grid {
      display: grid;
      grid-template-columns: 150px 1fr;
      gap: 12px 14px;
      align-items: center;
    }
    label { font-weight: 700; }
    input[type="file"], input[type="text"], input[type="number"], select {
      width: 100%;
      min-height: 38px;
      border: 1px solid var(--line);
      border-radius: 6px;
      padding: 8px 10px;
      font: inherit;
      background: #fff;
    }
    input[type="checkbox"] { width: 17px; height: 17px; }
    .inline { display: flex; gap: 10px; align-items: center; flex-wrap: wrap; }
    .inline input.short { max-width: 130px; }
    .hint { color: var(--muted); font-size: 14px; margin-top: 4px; }
    .toggle-row {
      display: flex;
      gap: 20px;
      align-items: center;
      flex-wrap: wrap;
      margin-bottom: 12px;
    }
    .toggle-row label {
      display: inline-flex;
      gap: 8px;
      align-items: center;
      font-weight: 700;
    }
    .targets {
      display: grid;
      gap: 10px;
    }
    .target-row {
      display: grid;
      grid-template-columns: minmax(220px, 1fr) 130px 42px;
      gap: 10px;
      align-items: center;
    }
    button {
      border: 1px solid var(--line);
      background: #fff;
      border-radius: 6px;
      padding: 10px 16px;
      font: inherit;
      cursor: pointer;
    }
    button.icon {
      width: 38px;
      height: 38px;
      padding: 0;
      font-size: 20px;
      line-height: 1;
    }
    button.primary {
      background: var(--accent);
      border-color: var(--accent);
      color: #fff;
      font-weight: 700;
    }
    button.primary:hover { background: var(--accent-dark); }
    button.link {
      border-color: #bfd0ff;
      color: var(--blue);
      background: #f7f9ff;
    }
    button:disabled { opacity: .55; cursor: not-allowed; }
    .actions {
      display: flex;
      gap: 10px;
      justify-content: flex-end;
      flex-wrap: wrap;
      border-top: 1px solid var(--line);
      padding-top: 16px;
      margin-top: 16px;
    }
    .status {
      margin-top: 14px;
      padding: 12px 14px;
      border: 1px solid var(--line);
      border-radius: 6px;
      background: var(--soft);
      white-space: pre-wrap;
    }
    .status.error {
      border-color: #f0b8b2;
      background: #fff4f2;
      color: var(--danger);
    }
    .log {
      margin-top: 12px;
      height: 140px;
      overflow: auto;
      border: 1px solid var(--line);
      border-radius: 6px;
      background: #fbfcfe;
      padding: 10px;
      font-family: Consolas, "Microsoft JhengHei UI", monospace;
      font-size: 13px;
      white-space: pre-wrap;
    }
    code {
      background: #edf2f7;
      border: 1px solid #d8dee8;
      padding: 1px 5px;
      border-radius: 4px;
    }
    @media (max-width: 760px) {
      header { padding-left: 18px; padding-right: 18px; }
      .shell { padding: 16px; }
      .grid { grid-template-columns: 1fr; }
      .target-row { grid-template-columns: 1fr; }
      .actions { justify-content: stretch; }
      button { flex: 1 1 auto; }
    }
  </style>
</head>
<body>
  <header>
    <h1>Excel Checklist 產生工具</h1>
    <div class="tabs">
      <button class="tab active" data-tab="run">設定與執行</button>
      <button class="tab" data-tab="guide">使用說明</button>
    </div>
  </header>

  <main>
    <section id="run" class="page active">
      <div class="shell">
        <h2>設定與執行</h2>

        <div class="config-block">
          <div class="block-title">
            <div>
              <h3>來源清單資料</h3>
              <p>選擇要讀取的清單，設定名稱在哪個工作表與欄位。</p>
            </div>
          </div>
          <div class="grid">
            <label for="sourceFile">來源清單</label>
            <div>
              <input id="sourceFile" type="file" accept=".xlsx">
              <div class="hint">含有議題名稱清單的 Excel。</div>
            </div>

            <label for="sourceSheet">來源工作表</label>
            <select id="sourceSheet"></select>

            <label for="sourceColumn">名稱欄位 / 起始列</label>
            <div class="inline">
              <input id="sourceColumn" class="short" type="text" value="E" maxlength="3" aria-label="名稱欄位">
              <input id="startRow" class="short" type="number" value="2" min="1" aria-label="起始列">
              <span class="hint">例：E 欄，第 2 列開始。</span>
            </div>
          </div>
        </div>

        <div class="config-block">
          <div class="block-title">
            <div>
              <h3>範本檔案</h3>
              <p>選擇要複製的 Checklist 範本，並設定要替換哪些位置。</p>
            </div>
          </div>
          <div class="grid">
            <label for="templateFile">範本檔案</label>
            <div>
              <input id="templateFile" type="file" accept=".xlsx">
              <div class="hint">工具會以此範本批次建立新的 Excel。</div>
            </div>
          </div>
        </div>

        <div class="config-block">
          <div class="block-title">
            <div>
              <h3>替換設定</h3>
              <p>可以替換範本內容、檔名，或兩者都替換。範本內容可設定多個工作表與儲存格。</p>
            </div>
          </div>

          <div class="toggle-row">
            <label><input id="replaceWorkbook" type="checkbox" checked> 替換範本內容</label>
            <label><input id="replaceFilename" type="checkbox" checked> 替換檔名</label>
          </div>

          <div class="grid">
            <label for="placeholder">替換文字</label>
            <input id="placeholder" type="text" value="BX-XXX_FSD議題名稱">
          </div>

          <div style="margin-top: 14px;">
            <label>範本內容替換位置</label>
            <div id="targets" class="targets" style="margin-top: 8px;"></div>
            <button id="addTarget" class="link" type="button" style="margin-top: 10px;">新增替換位置</button>
            <div class="hint">每一組都是「範本工作表 + 儲存格」。常用預設是 <code>檢核表 / A1</code> 或 <code>FSD Checklist / A1</code>。</div>
          </div>
        </div>

        <div class="actions">
          <button id="preview" type="button">預覽筆數</button>
          <button id="runButton" type="button" class="primary">產生並下載 zip</button>
        </div>

        <div id="status" class="status">請選擇來源清單與範本檔案。</div>
        <div id="log" class="log"></div>
      </div>
    </section>

    <section id="guide" class="page">
      <div class="shell">
        <h2>使用說明</h2>
        <h3>用途</h3>
        <p>此工具會讀取來源清單 Excel 的名稱資料，把每一筆名稱套進 Checklist 範本，最後下載一個 zip。zip 裡面的輸出資料夾會自動加日期時間，避免覆蓋之前產出。</p>

        <h3>使用流程</h3>
        <ul>
          <li>選擇來源清單，設定來源工作表、名稱欄位、起始列。</li>
          <li>選擇範本檔案。</li>
          <li>設定要替換的範本位置，例如 <code>檢核表 / A1</code>。</li>
          <li>確認是否要替換檔名。</li>
          <li>按「預覽筆數」確認資料，再按「產生並下載 zip」。</li>
        </ul>

        <h3>常用設定</h3>
        <ul>
          <li>模板一：來源工作表 <code>CR FSD (模板一)</code>，名稱欄位 <code>E</code>，範本工作表 <code>FSD Checklist</code>。</li>
          <li>模板二：來源工作表 <code>CR UTFT (模板二)</code>，名稱欄位 <code>E</code>，範本工作表 <code>檢核表</code>。</li>
        </ul>

        <h3>範本要求</h3>
        <ul>
          <li>要被替換的位置需包含 <code>BX-XXX_FSD議題名稱</code>，可在「替換文字」調整。</li>
          <li>若勾選替換檔名，檔名中的 <code>BX-XXX FSD議題名稱</code> 或 <code>BX-XXX_FSD議題名稱</code> 會自動換成清單名稱。</li>
        </ul>
      </div>
    </section>
  </main>

  <script>
    const $ = (id) => document.getElementById(id);
    const statusBox = $("status");
    const logBox = $("log");
    let templateSheets = [];

    document.querySelectorAll(".tab").forEach((tab) => {
      tab.addEventListener("click", () => {
        document.querySelectorAll(".tab").forEach((item) => item.classList.remove("active"));
        document.querySelectorAll(".page").forEach((item) => item.classList.remove("active"));
        tab.classList.add("active");
        $(tab.dataset.tab).classList.add("active");
      });
    });

    function setStatus(message, isError = false) {
      statusBox.textContent = message;
      statusBox.classList.toggle("error", isError);
    }

    function log(message) {
      logBox.textContent += message + "\n";
      logBox.scrollTop = logBox.scrollHeight;
    }

    function setSelectOptions(select, values) {
      select.innerHTML = "";
      values.forEach((value) => {
        const option = document.createElement("option");
        option.value = value;
        option.textContent = value;
        select.appendChild(option);
      });
    }

    function selectOptionByText(select, text) {
      for (const option of select.options) {
        if (option.textContent === text) {
          select.value = option.value;
          return true;
        }
      }
      return false;
    }

    function addTarget(sheet = "", cell = "A1") {
      const row = document.createElement("div");
      row.className = "target-row";

      const select = document.createElement("select");
      setSelectOptions(select, templateSheets);
      if (sheet) selectOptionByText(select, sheet);

      const input = document.createElement("input");
      input.type = "text";
      input.value = cell || "A1";
      input.placeholder = "A1";

      const remove = document.createElement("button");
      remove.type = "button";
      remove.className = "icon";
      remove.textContent = "x";
      remove.title = "移除此替換位置";
      remove.addEventListener("click", () => row.remove());

      row.append(select, input, remove);
      $("targets").appendChild(row);
    }

    function refreshTargetSheetOptions() {
      document.querySelectorAll("#targets select").forEach((select) => {
        const oldValue = select.value;
        setSelectOptions(select, templateSheets);
        if (oldValue) selectOptionByText(select, oldValue);
      });
    }

    async function fileToBase64(file) {
      const buffer = await file.arrayBuffer();
      let binary = "";
      const bytes = new Uint8Array(buffer);
      const chunkSize = 0x8000;
      for (let i = 0; i < bytes.length; i += chunkSize) {
        binary += String.fromCharCode.apply(null, bytes.subarray(i, i + chunkSize));
      }
      return btoa(binary);
    }

    async function postJson(url, payload) {
      const response = await fetch(url, {
        method: "POST",
        headers: { "Content-Type": "application/json" },
        body: JSON.stringify(payload),
      });
      if (!response.ok) {
        const text = await response.text();
        throw new Error(text || "執行失敗");
      }
      return response;
    }

    async function loadSheets(fileInput, select, kind) {
      const file = fileInput.files[0];
      select.innerHTML = "";
      if (!file) return;
      try {
        setStatus("讀取工作表中...");
        const response = await postJson("/api/sheets", {
          filename: file.name,
          data: await fileToBase64(file),
        });
        const data = await response.json();
        setSelectOptions(select, data.sheets);
        if (kind === "template") {
          templateSheets = data.sheets;
          refreshTargetSheetOptions();
          if (!document.querySelector("#targets .target-row")) addTarget(data.sheets[0] || "", "A1");
        }
        setStatus(`已讀取 ${file.name} 的工作表。`);
      } catch (error) {
        setStatus(error.message, true);
      }
    }

    function collectOptions() {
      const source = $("sourceFile").files[0];
      const template = $("templateFile").files[0];
      if (!source) throw new Error("請選擇來源清單 Excel。");
      if (!template) throw new Error("請選擇範本 Excel。");
      return { source, template };
    }

    function collectTargets() {
      const targets = [];
      document.querySelectorAll("#targets .target-row").forEach((row) => {
        const select = row.querySelector("select");
        const input = row.querySelector("input");
        if (select.value && input.value.trim()) {
          targets.push({ sheetName: select.value, cell: input.value.trim() });
        }
      });
      return targets;
    }

    async function buildPayload() {
      const { source, template } = collectOptions();
      return {
        source: { filename: source.name, data: await fileToBase64(source) },
        template: { filename: template.name, data: await fileToBase64(template) },
        sourceSheet: $("sourceSheet").value,
        sourceColumn: $("sourceColumn").value,
        startRow: Number($("startRow").value),
        placeholder: $("placeholder").value,
        outputPrefix: "output_checklists",
        replaceWorkbook: $("replaceWorkbook").checked,
        replaceFilename: $("replaceFilename").checked,
        replacementTargets: collectTargets(),
      };
    }

    $("sourceFile").addEventListener("change", () => loadSheets($("sourceFile"), $("sourceSheet"), "source"));
    $("templateFile").addEventListener("change", () => loadSheets($("templateFile"), document.createElement("select"), "template"));
    $("addTarget").addEventListener("click", () => addTarget(templateSheets[0] || "", "A1"));

    $("replaceWorkbook").addEventListener("change", () => {
      $("targets").style.opacity = $("replaceWorkbook").checked ? "1" : ".45";
      $("addTarget").disabled = !$("replaceWorkbook").checked;
    });

    $("preview").addEventListener("click", async () => {
      try {
        const payload = await buildPayload();
        const response = await postJson("/api/preview", payload);
        const data = await response.json();
        setStatus(`預計產生 ${data.count} 份。\n前幾筆：${data.sample.join("、")}`);
        log(`預覽成功：${data.count} 份`);
      } catch (error) {
        setStatus(error.message, true);
      }
    });

    $("runButton").addEventListener("click", async () => {
      const button = $("runButton");
      try {
        button.disabled = true;
        setStatus("產生中，請稍候...");
        log("開始產生");
        const response = await postJson("/api/run", await buildPayload());
        const blob = await response.blob();
        const disposition = response.headers.get("Content-Disposition") || "";
        const match = disposition.match(/filename="([^"]+)"/);
        const filename = match ? match[1] : "output_checklists.zip";
        const url = URL.createObjectURL(blob);
        const link = document.createElement("a");
        link.href = url;
        link.download = filename;
        document.body.appendChild(link);
        link.click();
        link.remove();
        URL.revokeObjectURL(url);
        setStatus(`完成，已下載 ${filename}`);
        log(`完成下載：${filename}`);
      } catch (error) {
        setStatus(error.message, true);
      } finally {
        button.disabled = false;
      }
    });
  </script>
</body>
</html>
"""


def decode_upload(payload: dict, key: str, folder: Path) -> Path:
    item = payload[key]
    filename = sanitize_filename_part(item.get("filename") or f"{key}.xlsx")
    path = folder / filename
    path.write_bytes(base64.b64decode(item["data"]))
    return path


def sheets_from_bytes(data: str) -> list[str]:
    wb = load_workbook(io.BytesIO(base64.b64decode(data)), read_only=True, data_only=True)
    return list(wb.sheetnames)


def make_options(payload: dict, folder: Path) -> FillOptions:
    source_path = decode_upload(payload, "source", folder)
    template_path = decode_upload(payload, "template", folder)
    targets = [
        ReplacementTarget(item.get("sheetName", ""), item.get("cell", ""))
        for item in payload.get("replacementTargets", [])
    ]
    return FillOptions(
        source_file=source_path,
        template_file=template_path,
        source_sheet=payload.get("sourceSheet", ""),
        source_column=payload.get("sourceColumn", "E"),
        start_row=int(payload.get("startRow", 2)),
        placeholder=payload.get("placeholder", DEFAULT_PLACEHOLDER),
        output_base_dir=folder,
        output_prefix=payload.get("outputPrefix", "output_checklists"),
        replacement_targets=targets,
        replace_workbook=bool(payload.get("replaceWorkbook", True)),
        replace_filename=bool(payload.get("replaceFilename", True)),
    )


def zip_output_dir(output_dir: Path) -> bytes:
    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as archive:
        for file in output_dir.rglob("*"):
            if file.is_file():
                archive.write(file, file.relative_to(output_dir.parent))
    return buffer.getvalue()


class AppHandler(BaseHTTPRequestHandler):
    server_version = "ExcelTemplateFiller/3.0"

    def do_GET(self) -> None:
        path = urlparse(self.path).path
        if path == "/":
            self.send_bytes(HTML.encode("utf-8"), "text/html; charset=utf-8")
            return
        if path == "/shutdown":
            self.send_bytes(b"OK", "text/plain; charset=utf-8")
            threading.Thread(target=self.server.shutdown, daemon=True).start()
            return
        self.send_error(HTTPStatus.NOT_FOUND)

    def do_POST(self) -> None:
        try:
            length = int(self.headers.get("Content-Length", "0"))
            payload = json.loads(self.rfile.read(length).decode("utf-8"))
            path = urlparse(self.path).path

            if path == "/api/sheets":
                self.send_json({"sheets": sheets_from_bytes(payload["data"])})
                return

            temp_dir = Path(tempfile.mkdtemp(prefix="excel_template_filler_"))
            try:
                options = make_options(payload, temp_dir)
                if path == "/api/preview":
                    validate_options(options)
                    names = read_names(options.source_file, options.source_sheet, options.source_column, options.start_row)
                    self.send_json({"count": len(names), "sample": names[:5]})
                    return

                if path == "/api/run":
                    result = fill_templates(options)
                    zip_bytes = zip_output_dir(result.output_dir)
                    zip_name = result.output_dir.name + ".zip"
                    headers = {
                        "Content-Disposition": f'attachment; filename="{zip_name}"',
                        "Cache-Control": "no-store",
                        "X-Output-Folder": result.output_dir.name,
                    }
                    self.send_bytes(zip_bytes, "application/zip", headers=headers)
                    return

                self.send_error(HTTPStatus.NOT_FOUND)
            finally:
                shutil.rmtree(temp_dir, ignore_errors=True)
        except Exception as exc:
            self.send_error_text(str(exc))

    def log_message(self, format: str, *args) -> None:
        return

    def send_json(self, data: dict) -> None:
        self.send_bytes(json.dumps(data, ensure_ascii=False).encode("utf-8"), "application/json; charset=utf-8")

    def send_error_text(self, message: str) -> None:
        encoded = message.encode("utf-8")
        self.send_response(HTTPStatus.BAD_REQUEST)
        self.send_header("Content-Type", "text/plain; charset=utf-8")
        self.send_header("Content-Length", str(len(encoded)))
        self.end_headers()
        self.wfile.write(encoded)

    def send_bytes(self, data: bytes, content_type: str, headers: dict[str, str] | None = None) -> None:
        self.send_response(HTTPStatus.OK)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(data)))
        for key, value in (headers or {}).items():
            self.send_header(key, value)
        self.end_headers()
        self.wfile.write(data)


def run_self_test() -> None:
    source = EXAMPLES_DIR / "sample_source_list.xlsx"
    template = EXAMPLES_DIR / "sample_utft_template.xlsx"
    temp_dir = Path(tempfile.mkdtemp(prefix="excel_template_filler_"))
    try:
        source_sheets = load_workbook(source, read_only=True).sheetnames
        template_sheets = load_workbook(template, read_only=True).sheetnames
        options = FillOptions(
            source_file=source,
            template_file=template,
            source_sheet=source_sheets[1],
            source_column="E",
            start_row=2,
            output_base_dir=temp_dir,
            output_prefix="self_test",
            replacement_targets=[ReplacementTarget(template_sheets[0], "A1")],
            replace_workbook=True,
            replace_filename=True,
        )
        result = fill_templates(options)
        if len(result.created_files) < 1:
            raise RuntimeError("Self-test did not create any files.")
    finally:
        shutil.rmtree(temp_dir, ignore_errors=True)


def main() -> None:
    if "--self-test" in sys.argv:
        run_self_test()
        return

    server = ThreadingHTTPServer(("127.0.0.1", 0), AppHandler)
    host, port = server.server_address
    open_app_url(f"http://{host}:{port}/")
    try:
        server.serve_forever()
    finally:
        server.server_close()


if __name__ == "__main__":
    main()
